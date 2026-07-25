from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_CSV, OUTPUT_JSON, OUTPUT_XLSX
from .models import PartRow

HEADERS = ["assembly", "ref", "oem_part_number", "description", "qty", "status", "superseded", "source_url", "page_title", "notes"]


def _write_sheet(ws, records: list[dict]) -> None:
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, len(records) + 1)}"
    for record in records:
        ws.append([record.get(header, "") for header in HEADERS])
    for index, header in enumerate(HEADERS, 1):
        width = max(len(header) + 2, min(65, max([len(str(r.get(header, ""))) for r in records] + [0]) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width


def export(rows: list[PartRow], errors: list[dict]) -> None:
    records = [row.to_dict() for row in rows]
    OUTPUT_XLSX.parent.mkdir(exist_ok=True)
    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader(); writer.writerows(records)
    OUTPUT_JSON.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")

    book = Workbook(); master = book.active; master.title = "Master Database"
    _write_sheet(master, records)
    sections: dict[str, list[dict]] = defaultdict(list)
    for record in records: sections[record["assembly"]].append(record)
    for section, section_records in sections.items():
        sheet_name = section[:31]
        _write_sheet(book.create_sheet(sheet_name), section_records)
    error_sheet = book.create_sheet("Errors")
    error_sheet.append(["assembly", "source_url", "error"])
    for error in errors: error_sheet.append([error.get("assembly", ""), error.get("source_url", ""), error.get("error", "")])
    book.save(OUTPUT_XLSX)

